import re
import pdfplumber
from openpyxl import Workbook


def extract_text_from_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = ""
        for page in pdf.pages:
            text += page.extract_text()
    return text


def extract_data(text):
    buildings = re.findall("pattern_for_building_name", text)
    floor_numbers = re.findall("pattern_for_floor_number", text)
    public_spaces = re.findall("pattern_for_public_spaces", text)
    apartment_numbers = re.findall("pattern_for_apartment_number", text)
    spaces_in_apartments = re.findall("pattern_for_spaces_in_apartment", text)

    return {
        "buildings": buildings,
        "floor_numbers": floor_numbers,
        "public_spaces": public_spaces,
        "apartment_numbers": apartment_numbers,
        "spaces_in_apartments": spaces_in_apartments
    }


def write_data_to_excel(data, output_file):
    wb = Workbook()
    ws = wb.active

    # Write headers
    headers = ["Building", "Floor", "Public Space",
               "Apartment Number", "Space"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header

    # Write data
    row = 2
    for building in data["buildings"]:
        for floor in data["floor_numbers"]:
            for public_space in data["public_spaces"]:
                for apartment_number in data["apartment_numbers"]:
                    for space in data["spaces_in_apartments"]:
                        # Customize this part to write data in the required format
                        ws.cell(row=row, column=1).value = building
                        ws.cell(row=row, column=2).value = floor
                        ws.cell(row=row, column=3).value = public_space
                        ws.cell(row=row, column=4).value = apartment_number
                        ws.cell(row=row, column=5).value = space
                        row += 1

    wb.save(output_file)


def main():
    pdf_files = ["04-pohjaB1.pdf"]
    output_file = "output.xlsx"

    # Initialize a data structure to hold the extracted information
    extracted_data = {
        "buildings": [],
        "floor_numbers": [],
        "public_spaces": [],
        "apartment_numbers": [],
        "spaces_in_apartments": []
    }

    for pdf_file in pdf_files:
        text = extract_text_from_pdf(pdf_file)
        data = extract_data(text)

        # Merge the extracted data
        for key in extracted_data:
            extracted_data[key].extend(data[key])

    write_data_to_excel(extracted_data, output_file)


if __name__ == "__main__":
    main()

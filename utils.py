from keywords import building_keywords, floor_keywords, apartment_keywords, room_keywords
import re
import pdfplumber
import pytesseract
from wand.image import Image
from openpyxl import Workbook
import tempfile
from PyPDF2 import PdfReader
from PIL import Image as PILImage
import string


# Text preprocessing in a desperate attempt to do some cleanup for the data we get

import string


def preprocess_text(text):
    # Convert text to lowercase
    text = text.lower()

    # Remove punctuation
    # text = text.translate(str.maketrans("", "", string.punctuation))

    # Replace multiple whitespaces with a single whitespace
    # text = re.sub(r'\s+', ' ', text)

    return text

# Function to extract relevant data from the text using regular expressions


def extract_text_from_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = ""
        for page in pdf.pages:
            # Extract selectable text from the current page
            page_text = page.extract_text()
            if page_text:
                text += preprocess_text(page_text) + "\n"

            # Convert the page to an image and extract text using OCR with Tesseract
            img = page.to_image(resolution=400)

            # Convert to black and white to improve OCR accuracy
            with tempfile.NamedTemporaryFile(suffix=".png") as temp_img:
                img.save(temp_img.name, format="PNG")
                with Image(filename=temp_img.name) as wand_img:
                    wand_img.type = 'bilevel'
                    # Save wand image as PNG and open it with PIL
                    with tempfile.NamedTemporaryFile(suffix=".png") as pil_temp_img:
                        wand_img.save(filename=pil_temp_img.name)
                        pil_img = PILImage.open(pil_temp_img.name)
                        ocr_text = pytesseract.image_to_string(pil_img)
                text += preprocess_text(ocr_text) + "\n"

        # Extract annotations and form fields using PyPDF2
        with open(pdf_path, "rb") as f:
            reader = PdfReader(f)
            for page_num in range(len(reader.pages)):
                page = reader.pages[page_num]
                if "/Annots" in page:
                    for annot in page["/Annots"]:
                        annot_obj = annot.getObject()
                        if "/Contents" in annot_obj:
                            text += preprocess_text(annot_obj["/Contents"].replace(
                                "\n", " ").strip()) + "\n"

            acroform = reader.trailer.get("/AcroForm")
            if acroform:
                form_fields = acroform["/Fields"]
                for field in form_fields:
                    field_obj = field.getObject()
                    if "/V" in field_obj:
                        value = field_obj["/V"]
                        if value not in ["", "/Off"]:
                            text += preprocess_text(value.replace("\n",
                                                    " ").strip()) + "\n"

    return text


def extract_data(text, apartment_pattern, floor_range, special_floors):
    buildings_data = {}

    floor_min, floor_max = map(int, floor_range.split("-"))
    special_floors = special_floors.split(",")

    building_keyword_pattern = "|".join(building_keywords)
    floor_keyword_pattern = r"(\d+)(?:\s*\.?\s*|\.)(?:" + \
        "|".join(floor_keywords) + ")"
    apartment_keyword_pattern = apartment_pattern
    room_keyword_pattern = "|".join(room_keywords)

    building_matches = re.findall(
        fr"(?:{building_keyword_pattern})\s+(\w+)", text)

    for building in building_matches:
        if building not in buildings_data:
            buildings_data[building] = {}

        floor_matches = re.finditer(floor_keyword_pattern, text)
        for floor_match in floor_matches:
            floor = floor_match.group(1)

            # Check if the floor is within the range or is a special floor
            if floor.isdigit():
                if int(floor) < floor_min or int(floor) > floor_max:
                    continue
            elif floor not in special_floors:
                continue

            if floor not in buildings_data[building]:
                # initialize with an empty set
                buildings_data[building][floor] = {}

            apartment_matches = re.findall(
                fr"(?:{apartment_keyword_pattern})", text)
            for apartment in apartment_matches:
                if apartment not in buildings_data[building][floor]:
                    buildings_data[building][floor][apartment] = {}

                room_matches = re.finditer(
                    fr"(?:{room_keyword_pattern})\s*([\d.,]+)", text)
                for room_match in room_matches:
                    room_type = room_match.group(0).split()[0]
                    room_size = room_match.group(1)
                    buildings_data[building][floor][apartment][room_type] = room_size

    return buildings_data


def write_data_to_excel(data, output_file):
    wb = Workbook()
    ws = wb.active

    row = 1
    for building, floors in data.items():
        ws.cell(row=row, column=1).value = "Building"
        ws.cell(row=row, column=2).value = building
        row += 1
        for floor, apartments in floors.items():
            ws.cell(row=row, column=1).value = "Floor"
            ws.cell(row=row, column=2).value = f"Kerros {floor}"
            row += 1
            for apartment, rooms in apartments.items():
                ws.cell(row=row, column=1).value = "Apartment"
                ws.cell(row=row, column=3).value = apartment
                row += 1
                for room in rooms.keys():
                    ws.cell(row=row, column=1).value = "Room"
                    ws.cell(row=row, column=4).value = room
                    row += 1

    wb.save(output_file)

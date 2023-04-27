import re
import pdfplumber
import pytesseract
from wand.image import Image
from openpyxl import Workbook
import tempfile
from PIL import Image as PILImage

# Function to extract text from a PDF file


def extract_text_from_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = ""
        for page in pdf.pages:
            # Extract selectable text from the current page
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"

            # Convert the page to an image and extract text using OCR with Tesseract
            # You may adjust the resolution if needed
            img = page.to_image(resolution=300)

            # Convert to black and white to improve OCR accuracy
            with tempfile.NamedTemporaryFile(suffix=".png") as temp_img:
                img.save(temp_img.name, format="PNG")
                with Image(filename=temp_img.name) as wand_img:
                    wand_img.type = 'bilevel'
                    # Save wand image as PNG and open it with PIL
                    with tempfile.NamedTemporaryFile(suffix=".png") as pil_temp_img:
                        wand_img.save(filename=pil_temp_img.name)
                        pil_img = PILImage.open(pil_temp_img.name)
                        ocr_text = pytesseract.image_to_string(pil_img)
                text += ocr_text + "\n"
    return text


# Function to extract relevant data from the text using regular expressions


def extract_data(text):
    buildings_data = {}
    building_floor_pairs = re.findall(
        r"RAKENNUS\s+(\w+),\s+((?:\d+\.\s+)?KERROS|VESIKATTO)", text)
    for building, floor in building_floor_pairs:
        if building not in buildings_data:
            buildings_data[building] = {floor}
        else:
            buildings_data[building].add(floor)

    # Convert sets to lists for consistency and sort the floors
    for building, floors in buildings_data.items():
        buildings_data[building] = sorted(
            list(floors), key=lambda x: (x.isdigit() == False, x))

    return buildings_data


# Function to write the extracted data to an Excel file
def write_data_to_excel(data, output_file):
    wb = Workbook()
    ws = wb.active

    headers = ["Building", "Floor"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header

    row = 2
    for building, floors in data.items():
        ws.cell(row=row, column=1).value = building
        for i, floor in enumerate(floors):
            if i > 0:
                row += 1
            ws.cell(row=row, column=2).value = floor

        row += 1

    wb.save(output_file)


# Main function to run the script


def main():
    # Specify the PDF file and the output Excel file
    pdf_file = "04-pohjaB1.pdf"
    output_file = "output.xlsx"

    # Extract text from the PDF file
    text = extract_text_from_pdf(pdf_file)
    print(f"Extracted text from {pdf_file}:")
    print(text)
    print("\n" + "=" * 80 + "\n")

    # Extract data from the text
    data = extract_data(text)

    # Write the extracted data to an Excel file
    write_data_to_excel(data, output_file)


# Run the main function when the script is executed
if __name__ == "__main__":
    main()
